#!/usr/bin/env python
import sys
import os
import re
import json
import os.path
import glob2
import tempfile

import requests
from openpyxl import load_workbook
from collections import OrderedDict

usage = """\
usage: extract_to_json.py pdf_mask xlsx_stats output_dir

    pdf_mask - mask to find pdf files with info about licenses and perms
    xlsx_stats - path to xlsx file with stats about branches
    output_dir - path to dir to store extracted data in json files
"""

JSON_SETTINGS = dict(ensure_ascii=False, indent=4, sort_keys=True)


def slice_and_dice(text, sections):
    regex = '(' + ')|('.join(map(re.escape, sections)) + ')'
    return filter(None, map(str.strip, filter(None, re.split(regex, text))))


def get_sections_as_dict(text, sections):
    result = slice_and_dice(text, sections)

    res = {}
    prev_value = None

    for r in result:
        if r in sections:
            if prev_value is not None:
                res[prev_value] = ""
            prev_value = r
        else:
            res[prev_value] = r
            prev_value = None

    return res


def parse_numbered_list(lines):
    first_column_len = len(re.match("(\s*\w+\s*)", lines[0]).group(0))

    accum = []

    prev_code = None
    prev_desc = None

    for l in lines:
        code, desc = re.match("(\s*\w+\s*)(.*)$", l).groups()
        if len(code) > first_column_len:
            desc = code + desc
            code = ""

        code = code.strip()
        desc = desc.strip()

        if code:
            if prev_code:
                accum.append((prev_code, prev_desc))

            prev_code = code
            prev_desc = desc
        else:
            prev_desc += " " + desc

        if prev_code:
            accum.append((prev_code, prev_desc))

    return OrderedDict(accum)


class NBUParser(object):
    def __init__(self, xlsx_stats):
        super(NBUParser, self).__init__()
        self.accum = []
        self.bank_stats = {}
        self.xlsx_stats = xlsx_stats
        self.load_stats_spreadsheet()

    def load_stats_spreadsheet(self):
        wb = load_workbook(self.xlsx_stats, read_only=True)
        sheetnames = wb.sheetnames
        dates = ("01.01.2008", "01.01.2009", "01.01.2010", "01.01.2011",
                 "01.01.2012", "01.01.2013", "01.01.2014", "01.01.2015",
                 "01.07.2015", "01.08.2015")

        regions = (
            "Вінніцька", "Волинська", "Дніпропетровська", "Донецька",
            "Житомирська", "Закарпатська", "Запоріжська",
            "Івано-Франківська", "Київ", "Київська", "Кіровоградська",
            "Луганська", "Львівська", "Миколаївська", "Одеська",
            "Полтавська", "Рівненська", "Сумська", "Тернопільська",
            "Харківська", "Херсонська", "Хмельницька", "Черкаська",
            "Чернігівська", "Чернівецька"
        )

        ws = wb[sheetnames[0]]
        for i, r in enumerate(ws.rows):
            if i < 6:
                continue

            code, name, mfo = r[0].value, r[1].value, r[2].value

            if all([code, name, mfo]):
                values = map(lambda x: x.value if x.value else 0, r[3:13])

                by_date = [
                    {
                        "Дата": date,
                        "Кількість": value,
                    }
                    for date, value in zip(dates, values)]

                self.bank_stats[mfo] = {
                    "Код банку": code,
                    "Структурні підрозділи по датах": by_date
                }

        ws = wb[sheetnames[1]]
        for i, r in enumerate(ws.rows):
            if i < 6:
                continue

            code, name, mfo = r[0].value, r[1].value, r[2].value

            if all([code, name, mfo]):
                values = map(lambda x: x.value if x.value else 0,
                             r[5:5 + len(regions)])

                by_region = [
                    {
                        "Регіон": region,
                        "Кількість": value,
                    }
                    for region, value in zip(regions, values)]

                self.bank_stats[mfo][
                    "Структурні підрозділи по регіонах"] = by_region

    def parse_main_license(self, license):
        license_lines = license.splitlines()
        license_data = get_sections_as_dict(
            "номер ліцензії {0}".format(license_lines[0]),
            ["номер ліцензії", "від", "номер бланка"])

        license_data["Перелік операцій ліцензії"] = parse_numbered_list(
            license_lines[2:])

        return license_data

    def parse_permissions(self, permissions):
        permissions_lines = permissions.splitlines()

        permissions_data = get_sections_as_dict(
            "номер дозволу {0}".format(permissions_lines[0]),
            ["номер дозволу", "від"])

        permissions_data["Перелік операцій дозвола"] = parse_numbered_list(
            permissions_lines[2:])

        return permissions_data

    def expand_licenses(self, entry):
        entry["Назва банку"] = re.sub("\s+", " ", entry["Назва банку"])
        entry["Адреса"] = re.sub("\s+", " ", entry["Адреса"])

        if "Банківська ліцензія" in entry:
            lic_type = "Банківська ліцензія"
        elif "Ліцензія санаційного банку" in entry:
            lic_type = "Ліцензія санаційного банку"
            entry[lic_type] = entry["Ліцензія санаційного банку"]
            del entry["Ліцензія санаційного банку"]
        elif "Ліцензія санаційного банку" in entry:
            lic_type = "Ліцензія санаційного банку"
        else:
            return entry

        licenses = re.split("Дозвіл ", entry[lic_type])

        entry[lic_type] = self.parse_main_license(licenses[0])
        entry["Дозволи"] = list(map(self.parse_permissions, licenses[1:]))

        return entry

    def parse_ownership(self, mfo):
        url = "http://www.bank.gov.ua/files/Shareholders/{0}/".format(mfo)
        resp = requests.get(url)

        if resp.status_code == 200:
            return [
                {
                    "Дата": x[1],
                    "Посилання": url + x[0]
                } for x in re.findall(
                    "([\d_]*\.pdf).*(\d{4}\.\d{2}\.\d{2})", resp.text)]
        else:
            print("Ownership structure for {0} not found".format(mfo))

        return {}

    def convert_one(self, fname):
        output_file = tempfile.mktemp(prefix="yes_i_know_you_deprecated")

        """
        pdftotext - is installed poppler-utils package for PDF to text
        conversion;

        keys:
            -raw   Keep the text in content stream order;
            -nopgbrk  Don’t insert page breaks between  pages.
        """
        os.system('pdftotext -layout -nopgbrk "{0}" "{1}"'.format(
            fname, output_file))

        with open(output_file, 'r') as f:
            # converted text from pdf file
            text = f.read()

        os.unlink(output_file)

        entries = filter(str.strip, re.split(re.escape("Назва банку"), text))
        entry_sections = ["Назва банку", "Номер банку (МФО)", "Код ЄДРПОУ",
                          "Адреса", "Банківська ліцензія", "Дата відкликання",
                          "Підстави прийняття рішення",
                          "Ліцензія санаційного банку",
                          "Ліцензія санаційного банку"]

        expanded_entries = []

        try:
            expanded_entries = list(map(
                lambda x: self.expand_licenses(get_sections_as_dict(
                    "Назва банку {0}".format(x), entry_sections)),
                entries))
        except:
            print(fname)

        mfo, _ = os.path.splitext(os.path.basename(fname))
        mfo = int(mfo)

        res = {
            "МФО": mfo,
            "Ліцензії": list(expanded_entries),
            "Структура власності": self.parse_ownership(mfo)
        }

        if mfo in self.bank_stats:
            res.update(self.bank_stats[mfo])

        self.accum.append(res)

    def convert_many(self, mask):
        for fl in glob2.glob(mask):
            self.convert_one(fl)

    def export_to_dir(self, out_dir):
        with open(os.path.join(out_dir, "complete.json"), "w") as fp:
            json.dump(self.accum, fp, **JSON_SETTINGS)

        index = []

        for struct in self.accum:
            ind_fname = "{0}.json".format(struct["МФО"])
            index_entry = {
                "МФО": struct["МФО"],
                "Назва банку": struct["Ліцензії"][0]["Назва банку"],
                "Адреса": struct["Ліцензії"][0]["Адреса"],
                "Код банку": struct.get("Код банку", "-")
            }

            if "Структурні підрозділи по датах" in struct:
                index_entry.update({
                    "Філії": struct["Структурні підрозділи по датах"][-1]
                })

            if struct["Структура власності"]:
                index_entry.update({
                    "Структура власності": struct["Структура власності"][0]
                })
            index.append(index_entry)

            with open(os.path.join(out_dir, ind_fname), "w") as fp:
                json.dump(struct, fp, **JSON_SETTINGS)

        with open(os.path.join(out_dir, "index.json"), "w") as fp:
            json.dump(index, fp, **JSON_SETTINGS)


if __name__ == "__main__":
    if len(sys.argv) < 4:
        sys.exit(usage)

    parser = NBUParser(sys.argv[2])
    parser.convert_many(sys.argv[1])
    parser.export_to_dir(sys.argv[3])
